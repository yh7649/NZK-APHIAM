"""Build editable KEPCO fuel-technology EF handoff workbook."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from nzk_aphiam.config.paths import PROJECT_ROOT

TABLE_DIR = PROJECT_ROOT / "results" / "tables" / "kepco" / "annual_handoff"
OUTPUT_PATH = TABLE_DIR / "kepco_annual_ef_editable_handoff.xlsx"

INPUTS = {
    "National all years": TABLE_DIR / "kepco_annual_ef_editable_by_fuel_technology.csv",
    "Provincial all years": TABLE_DIR / "kepco_annual_ef_editable_by_province_fuel_technology.csv",
}

MACRO_YEARS = {"2021", "2025"}


def read_csv(path: Path) -> list[list[str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.reader(handle))


def filter_years(rows: list[list[str]], years: set[str]) -> list[list[str]]:
    if not rows:
        return rows
    header, body = rows[0], rows[1:]
    year_index = header.index("year")
    return [header] + [row for row in body if row[year_index] in years]


def write_sheet(workbook: Workbook, title: str, rows: list[list[str]]) -> None:
    sheet = workbook.create_sheet(title=title)
    for row in rows:
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_cells in sheet.columns:
        header = str(column_cells[0].value or "")
        width = 14
        if header in {"fuel_technology"}:
            width = 34
        elif header in {"plant_province"}:
            width = 18
        elif "kg_per_mwh" in header:
            width = 18
        elif header in {"plants_records", "generation_coverage", "months_covered"}:
            width = 20
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def build_workbook() -> Path:
    missing = [str(path) for path in INPUTS.values() if not path.exists()]
    if missing:
        raise FileNotFoundError(
            "Missing editable KEPCO handoff CSVs. Run the R analysis first:\n" + "\n".join(missing)
        )

    workbook = Workbook()
    readme = workbook.active
    readme.title = "README"
    readme.append(["KEPCO annual fuel x technology EF handoff"])
    readme.append(["Pollutant cells show weighted EF, monthly median, and [p10, p90]."])
    readme.append(["Plants / records is unique plant sites / valid source records."])
    readme.append(["MACRO sheets retain only 2021 and 2025. All-years sheets are raw outputs."])
    readme["A1"].font = Font(bold=True, size=14)
    readme.column_dimensions["A"].width = 90

    national = read_csv(INPUTS["National all years"])
    provincial = read_csv(INPUTS["Provincial all years"])
    write_sheet(workbook, "national_all_years", national)
    write_sheet(workbook, "provincial_all_years", provincial)
    write_sheet(workbook, "macro_2021_2025_national", filter_years(national, MACRO_YEARS))
    write_sheet(
        workbook,
        "macro_2021_2025_provincial",
        filter_years(provincial, MACRO_YEARS),
    )

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)
    return OUTPUT_PATH


def main() -> None:
    print(build_workbook())


if __name__ == "__main__":
    main()
